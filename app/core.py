"""照片批次加註文字 — 純邏輯層。

不依賴 Tkinter,負責:
  • 讀取照片資訊(檔名 / 拍攝日期時間 / GPS 座標,來自 EXIF)
  • 載入對照表(csv / xlsx)並依檔名查對應文字
  • 字型清單與載入(快取)
  • 文字樣式 / 預設樣式
  • 把文字繪到照片四角之一(支援白字黑框等描邊效果)
  • 產生輸出檔名 + 存檔
"""
from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field
from pathlib import Path

from PIL import ExifTags, Image, ImageDraw, ImageFont, ImageOps

# 支援的影像副檔名
IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp")


# ---------------------------------------------------------------------------
# 字型
# ---------------------------------------------------------------------------

FONTS_DIR = Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts"

# 精選對中文友善的常見 Windows 字型: (顯示名稱, 檔名, ttc 內索引)
_CANDIDATE_FONTS = [
    ("微軟正黑體", "msjh.ttc", 0),
    ("微軟正黑體 粗體", "msjhbd.ttc", 0),
    ("微軟正黑體 細", "msjhl.ttc", 0),
    ("微軟雅黑", "msyh.ttc", 0),
    ("微軟雅黑 粗體", "msyhbd.ttc", 0),
    ("新細明體", "mingliu.ttc", 0),
    ("標楷體", "kaiu.ttf", 0),
    ("黑體", "simhei.ttf", 0),
    ("Arial", "arial.ttf", 0),
    ("Arial 粗體", "arialbd.ttf", 0),
    ("Times New Roman", "times.ttf", 0),
    ("Consolas", "consola.ttf", 0),
    ("Segoe UI", "segoeui.ttf", 0),
]


def available_fonts() -> list[tuple[str, str, int]]:
    """回傳實際存在於系統字型資料夾的 (顯示名稱, 絕對路徑, 索引) 清單。"""
    out: list[tuple[str, str, int]] = []
    for name, fn, idx in _CANDIDATE_FONTS:
        p = FONTS_DIR / fn
        if p.exists():
            out.append((name, str(p), idx))
    return out


_font_cache: dict[tuple[str, int, int], ImageFont.FreeTypeFont] = {}


def load_font(path: str, index: int, size: int) -> ImageFont.FreeTypeFont:
    key = (path, index, size)
    f = _font_cache.get(key)
    if f is None:
        f = ImageFont.truetype(path, size=size, index=index)
        _font_cache[key] = f
    return f


# ---------------------------------------------------------------------------
# 預設樣式 / 位置
# ---------------------------------------------------------------------------

# (顯示名稱) -> (填色, 描邊色 或 None)
PRESETS: dict[str, tuple[str, str | None]] = {
    "白字黑框": ("#FFFFFF", "#000000"),
    "黑字白框": ("#000000", "#FFFFFF"),
    "黃字黑框": ("#FFFF00", "#000000"),
    "紅字白框": ("#FF0000", "#FFFFFF"),
    "純白字": ("#FFFFFF", None),
    "純黑字": ("#000000", None),
    "自訂": None,  # type: ignore[dict-item]
}
PRESET_NAMES = list(PRESETS.keys())

POSITIONS = ["左上", "右上", "左下", "右下"]


@dataclass
class TextStyle:
    font_path: str
    font_index: int = 0
    preset: str = "白字黑框"
    fill: str = "#FFFFFF"
    stroke: str | None = "#000000"
    stroke_ratio: float = 0.12     # 描邊粗細相對字級的比例
    size_mode: str = "auto"        # "auto"(依圖寬%) | "fixed"(固定像素)
    size_percent: float = 4.0      # auto 時:字級 = 圖寬 * 此 %
    size_px: int = 48              # fixed 時:固定字級
    position: str = "左下"
    margin_percent: float = 3.0    # 邊距 = 圖寬 * 此 %
    line_spacing_ratio: float = 0.2


def font_px_for_width(style: TextStyle, img_w: int) -> int:
    if style.size_mode == "auto":
        return max(8, round(img_w * style.size_percent / 100.0))
    return max(4, int(style.size_px))


def annotate_image(img: Image.Image, text: str, style: TextStyle) -> Image.Image:
    """把 text 畫到 img 上指定角落,原地修改並回傳(同一物件)。"""
    if not text:
        return img
    if img.mode not in ("RGB", "RGBA"):
        img = img.convert("RGB")
    w, h = img.size
    px = font_px_for_width(style, w)
    font = load_font(style.font_path, style.font_index, px)
    stroke_w = max(1, round(px * style.stroke_ratio)) if style.stroke else 0
    spacing = round(px * style.line_spacing_ratio)
    margin = round(w * style.margin_percent / 100.0)
    align = "left" if "左" in style.position else "right"

    draw = ImageDraw.Draw(img)
    bbox = draw.multiline_textbbox(
        (0, 0), text, font=font, stroke_width=stroke_w, spacing=spacing, align=align
    )
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]

    if "左" in style.position:
        x = margin - bbox[0]
    else:
        x = w - margin - tw - bbox[0]
    if "上" in style.position:
        y = margin - bbox[1]
    else:
        y = h - margin - th - bbox[1]

    draw.multiline_text(
        (x, y), text, font=font, fill=style.fill,
        stroke_width=stroke_w, stroke_fill=style.stroke or None,
        spacing=spacing, align=align,
    )
    return img


# ---------------------------------------------------------------------------
# 照片資訊 (EXIF)
# ---------------------------------------------------------------------------

_TAG_DATETIME_ORIGINAL = 36867   # DateTimeOriginal
_TAG_DATETIME = 306              # DateTime
_TAG_GPSINFO = 0x8825

# 可供「讀取照片資訊」模式選取的欄位: (顯示名稱, 內部 key)
META_FIELDS = [
    ("檔名(去副檔名)", "filename"),
    ("檔名(含副檔名)", "filename_ext"),
    ("拍攝日期時間", "datetime"),
    ("拍攝日期", "date"),
    ("拍攝時間", "time"),
    ("GPS座標", "gps"),
    ("緯度", "lat"),
    ("經度", "lon"),
]
META_LABEL_OF = {k: lbl for lbl, k in META_FIELDS}


def _gps_to_deg(value, ref) -> float | None:
    try:
        d, m, s = value
        deg = float(d) + float(m) / 60.0 + float(s) / 3600.0
    except Exception:
        return None
    if str(ref).upper().strip() in ("S", "W"):
        deg = -deg
    return deg


def read_photo_info(path: str) -> dict:
    """讀照片基本資訊。EXIF 缺失時對應欄位留空,不會丟例外。"""
    p = Path(path)
    info = {
        "filename": p.stem,
        "filename_ext": p.name,
        "datetime": "",
        "date": "",
        "time": "",
        "lat": None,
        "lon": None,
        "gps": "",
    }
    try:
        with Image.open(path) as im:
            exif = im.getexif()
    except Exception:
        return info
    if not exif:
        return info

    dt = exif.get(_TAG_DATETIME_ORIGINAL) or exif.get(_TAG_DATETIME)
    if dt:
        s = str(dt).strip()
        parts = s.split(" ", 1)
        date_part = parts[0].replace(":", "-")
        time_part = parts[1].strip() if len(parts) > 1 else ""
        info["date"] = date_part
        info["time"] = time_part
        info["datetime"] = (date_part + " " + time_part).strip()

    try:
        gps = exif.get_ifd(_TAG_GPSINFO)
    except Exception:
        gps = None
    if gps:
        lat = _gps_to_deg(gps.get(2), gps.get(1))
        lon = _gps_to_deg(gps.get(4), gps.get(3))
        if lat is not None and lon is not None:
            info["lat"] = lat
            info["lon"] = lon
            info["gps"] = f"{lat:.6f}, {lon:.6f}"
    return info


def text_from_meta(info: dict, fields: list[str], sep: str = "\n") -> str:
    """依選取的欄位順序組合文字;空值欄位略過。"""
    parts: list[str] = []
    for key in fields:
        v = info.get(key)
        if v is None or v == "":
            continue
        if isinstance(v, float):
            v = f"{v:.6f}"
        parts.append(str(v))
    return sep.join(parts)


# ---------------------------------------------------------------------------
# 對照表 (csv / xlsx)
# ---------------------------------------------------------------------------

def list_sheets(path: str) -> list[str]:
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    return []


def read_table(path: str, sheet: str | None = None) -> tuple[list[str], list[list]]:
    """讀對照表,回傳 (表頭, 資料列)。支援 csv / xlsx。"""
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(path, sheet)
    if ext == ".csv":
        return _read_csv(path)
    raise ValueError("對照表僅支援 .csv / .xlsx")


def _read_xlsx(path: str, sheet: str | None) -> tuple[list[str], list[list]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            return [], []
        headers = [("" if h is None else str(h)).strip() for h in header]
        rows = [list(r) for r in rows_iter]
        return headers, rows
    finally:
        wb.close()


def _read_csv(path: str) -> tuple[list[str], list[list]]:
    last_err: Exception | None = None
    for enc in ("utf-8-sig", "cp950", "big5", "utf-8", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)
            if not rows:
                return [], []
            headers = [c.strip() for c in rows[0]]
            return headers, [r for r in rows[1:]]
        except (UnicodeDecodeError, UnicodeError) as e:
            last_err = e
            continue
    raise last_err or ValueError("無法解析 CSV 編碼")


def _norm(s) -> str:
    return str(s).strip().lower()


def build_mapping(headers: list[str], rows: list[list], name_col: str, text_col: str) -> dict[str, str]:
    """以「照片名稱」欄為 key、文字欄為值建索引。

    同時用「含副檔名」與「去副檔名」兩種形式當 key,查詢時兩種都比對。
    """
    ni = headers.index(name_col)
    ti = headers.index(text_col)
    m: dict[str, str] = {}
    for r in rows:
        name = "" if ni >= len(r) or r[ni] is None else str(r[ni]).strip()
        if not name:
            continue
        text = "" if ti >= len(r) or r[ti] is None else str(r[ti])
        m.setdefault(_norm(name), text)
        m.setdefault(_norm(Path(name).stem), text)
    return m


def lookup_text(mapping: dict[str, str], photo_path: str) -> str | None:
    p = Path(photo_path)
    for key in (p.name, p.stem):
        t = mapping.get(_norm(key))
        if t is not None:
            return t
    return None


# ---------------------------------------------------------------------------
# 輸出檔名 / 存檔
# ---------------------------------------------------------------------------

def output_filename(src_path: str, naming_mode: str, prefix: str, index: int, digits: int) -> str:
    """naming_mode: "same"(同原檔名) | "custom"(自訂前綴 + 流水號)。保留原副檔名。"""
    p = Path(src_path)
    if naming_mode == "custom":
        stem = f"{prefix}{str(index).zfill(max(1, digits))}"
    else:
        stem = p.stem
    return stem + p.suffix


def save_image(img: Image.Image, out_path: str) -> None:
    ext = Path(out_path).suffix.lower()
    if ext in (".jpg", ".jpeg"):
        if img.mode != "RGB":
            img = img.convert("RGB")
        img.save(out_path, quality=95)
    elif ext == ".webp":
        img.save(out_path, quality=95)
    else:
        if ext in (".bmp",) and img.mode == "RGBA":
            img = img.convert("RGB")
        img.save(out_path)


def load_for_annotation(src_path: str) -> Image.Image:
    """開圖並依 EXIF 方向轉正(手機直拍照片才不會文字貼錯角)。"""
    with Image.open(src_path) as im:
        im = ImageOps.exif_transpose(im)
        return im.copy()
